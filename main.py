"""
Sistema de Gerenciamento dos Reparos — Mobile (Flet 0.24.1)
Funcionalidades completas: lista, filtros, formulário, fotos, materiais,
histórico, exportação Excel, configuração do banco.
"""
import flet as ft
import sqlite3, hashlib, json, os, io, threading, shutil
from datetime import datetime
from pathlib import Path

# ── Constantes ────────────────────────────────────────────────────────────────
ADMIN_PASSWORD = "#Val1001"
APP_DIR  = Path(__file__).parent
CFG_FILE = APP_DIR / "config.json"

CAT_SUBCAT = {
    "Calçada":["Reparo e/ou adequação de calçada","Reparo e/ou adequação de piso intertravado"],
    "Drenagem":["Reparo e/ou adequação de Boca de Lobo","Reparo e/ou adequação de caixa ralo",
        "Reparo e/ou adequação de poço de visita (PV)","Limpeza e desobstrução de rede",
        "Reparo e/ou adequação de rede de drenagem","Reparo e/ou adequação de caixa coletora",
        "Reparo e/ou adequação de tampa de concreto","Travessia de drenagem"],
    "Manutenções e Reparos específicos":["Manutenções Gerais","Manutenção de guarda-corpo",
        "Reparo de estruturas de madeira","Travessia para pedestres",
        "Reparo de estruturas de concreto","Reparo de abrigo de passageiros",
        "Fresagem de vias","Reparo de telhado"],
    "Vias públicas":["Tapa Buraco Emergencial (asfalto frio)","Manutenção de vias",
        "Reparo de via com paralelepípedos","Reparo de base para pavimentação asfáltica"],
    "Meio-fio":["Reparo e/ou substituição de meio-fio"],
    "Muro e Mureta":["Reparo de muro e/ou mureta"],
    "Base de concreto":["Reparo de piso ou base de concreto"],
    "Retirada de Material":["Retirada de resíduos de obra"],
    "Rios e Canais":["Desobstrução de valas e córregos","Limpeza de corpos hídricos"],
    "Movimentação de solo":["Troca de solo","Reconstituição de talude","Reconstituição de solo"],
}
SITUACOES = ["Aberto","Em execução","Cancelado","Finalizado"]
NUCLEOS   = ["CENTRO I","CENTRO II","CORDEIRINHO","ESPRAIADO","INOÃ","SÃO JOSÉ",
             "ITAIPUAÇU I","ITAIPUAÇU II"]
NUCLEO_DISTRITO = {
    "CENTRO I":"1º DISTRITO - CENTRO","CENTRO II":"1º DISTRITO - CENTRO",
    "CORDEIRINHO":"2º DISTRITO - PONTA NEGRA","ESPRAIADO":"2º DISTRITO - PONTA NEGRA",
    "INOÃ":"3º DISTRITO - INOÃ","SÃO JOSÉ":"3º DISTRITO - INOÃ",
    "ITAIPUAÇU I":"4º DISTRITO - ITAIPUAÇU","ITAIPUAÇU II":"4º DISTRITO - ITAIPUAÇU",
}
DISTRITO_BAIRROS = {
    "CENTRO I":["Centro","Mumbuca","Araçatiba","Flamengo","Barra de Maricá"],
    "CENTRO II":["Centro","Mumbuca","Araçatiba","Flamengo","Barra de Maricá"],
    "CORDEIRINHO":["Cordeirinho","Ponta Negra","Bambuí","Guaratiba","Bananal"],
    "ESPRAIADO":["Espraiado"],
    "INOÃ":["Inoã","São José"],"SÃO JOSÉ":["Inoã","São José"],
    "ITAIPUAÇU I":["Itaipuaçu","Recanto","Jardim Atlântico","Barroco"],
    "ITAIPUAÇU II":["Itaipuaçu","Recanto","Jardim Atlântico","Barroco"],
}

# Cores
BG="#0d0d1a"; BG2="#12121f"; BG3="#16162e"; CARD="#1a1a38"
BORDER="#2e2e55"; ACCENT="#3d9be9"; TEXT="#d0d8f0"; TEXT2="#9aabcc"
TEXT3="#667799"; SUCCESS="#1a6b3c"; DANGER="#b53030"; WHITE="#ffffff"
SIT_COR = {
    "ABERTO":("#3a3000","#ffdd66"), "EM EXECUÇÃO":("#0d2a4a","#88ccff"),
    "CANCELADO":("#3a0d0d","#ff8888"), "FINALIZADO":("#0d3a1a","#7aff9a"),
}

# ── Config ────────────────────────────────────────────────────────────────────
_config = {}

def load_config():
    global _config
    if CFG_FILE.exists():
        try:
            _config = json.loads(CFG_FILE.read_text("utf-8"))
        except Exception:
            _config = {}
    return _config

def save_config(db_path: str):
    global _config
    foto_dir = str(Path(db_path).parent / "fotos")
    _config = {"db_path": db_path, "foto_dir": foto_dir}
    Path(foto_dir).mkdir(parents=True, exist_ok=True)
    CFG_FILE.write_text(json.dumps(_config, indent=2, ensure_ascii=False), "utf-8")

def get_db_path() -> str | None:
    return _config.get("db_path")

def get_foto_dir() -> str:
    return _config.get("foto_dir", str(APP_DIR / "fotos"))

def has_config() -> bool:
    p = get_db_path()
    return bool(p and Path(p).parent.exists())

# ── Database ──────────────────────────────────────────────────────────────────
def get_conn():
    conn = sqlite3.connect(get_db_path(), timeout=10)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")
    return conn

def init_db():
    p = get_db_path()
    if not p: return False
    try:
        Path(p).parent.mkdir(parents=True, exist_ok=True)
        conn = sqlite3.connect(p, timeout=10)
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS solicitacoes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                data_abertura TEXT NOT NULL, solicitacao TEXT NOT NULL UNIQUE,
                nucleo TEXT, solicitante TEXT, endereco TEXT, bairro TEXT,
                distrito TEXT, categoria TEXT, subcategoria TEXT, encarregado TEXT,
                situacao TEXT, data_inicio TEXT, data_conclusao TEXT,
                observacao TEXT, localizacao TEXT,
                criado_em TEXT DEFAULT (datetime('now','localtime')),
                atualizado_em TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS fotos (
                id INTEGER PRIMARY KEY AUTOINCREMENT, solicitacao_id INTEGER NOT NULL,
                nome_arquivo TEXT NOT NULL, descricao TEXT,
                adicionada_em TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (solicitacao_id) REFERENCES solicitacoes(id) ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS materiais (
                id INTEGER PRIMARY KEY AUTOINCREMENT, solicitacao_id INTEGER NOT NULL,
                nome TEXT NOT NULL, quantidade TEXT, tipo TEXT,
                FOREIGN KEY (solicitacao_id) REFERENCES solicitacoes(id) ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS usuarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT NOT NULL,
                sobrenome TEXT NOT NULL, login TEXT NOT NULL UNIQUE,
                senha_hash TEXT NOT NULL, criado_em TEXT DEFAULT (datetime('now','localtime'))
            );
            CREATE TABLE IF NOT EXISTS historico (
                id INTEGER PRIMARY KEY AUTOINCREMENT, solicitacao_id INTEGER,
                solicitacao_num TEXT NOT NULL, usuario TEXT NOT NULL,
                data_hora TEXT NOT NULL, campo TEXT NOT NULL,
                valor_anterior TEXT, valor_novo TEXT
            );
            CREATE TABLE IF NOT EXISTS excluidos (
                id INTEGER PRIMARY KEY AUTOINCREMENT, solicitacao TEXT NOT NULL,
                excluido_por TEXT, excluido_em TEXT NOT NULL, dados_json TEXT
            );
        """)
        conn.commit(); conn.close()
        return True
    except Exception as e:
        print("init_db:", e); return False

def hash_senha(s): return hashlib.sha256(s.encode()).hexdigest()

def autenticar(login, senha):
    try:
        conn = get_conn()
        row = conn.execute(
            "SELECT nome,sobrenome FROM usuarios WHERE login=? AND senha_hash=?",
            (login.strip(), hash_senha(senha))
        ).fetchone()
        conn.close()
        return dict(row) if row else None
    except Exception: return None

def cadastrar_usuario(nome, sobrenome, login, senha):
    try:
        conn = get_conn()
        conn.execute("INSERT INTO usuarios (nome,sobrenome,login,senha_hash) VALUES (?,?,?,?)",
                     (nome.strip(), sobrenome.strip(), login.strip(), hash_senha(senha)))
        conn.commit(); conn.close(); return True, "Usuário cadastrado!"
    except sqlite3.IntegrityError: return False, "Login já em uso."
    except Exception as e: return False, str(e)

def gerar_proxima():
    ano = datetime.now().year
    try:
        conn = get_conn()
        row = conn.execute(
            "SELECT solicitacao FROM solicitacoes WHERE solicitacao LIKE ? "
            "ORDER BY CAST(substr(solicitacao,1,5) AS INTEGER) DESC LIMIT 1",
            (f"%/{ano}",)
        ).fetchone()
        conn.close()
        num = int(row["solicitacao"].split("/")[0]) + 1 if row else 1
        return f"{num:05d}/{ano}"
    except: return f"00001/{ano}"

def salvar_historico(sol_id, sol_num, usuario, diff):
    if not diff: return
    try:
        conn = get_conn()
        now = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        for campo, antes, depois in diff:
            conn.execute(
                "INSERT INTO historico (solicitacao_id,solicitacao_num,usuario,"
                "data_hora,campo,valor_anterior,valor_novo) VALUES (?,?,?,?,?,?,?)",
                (sol_id, sol_num, usuario, now, campo, antes or "", depois or "")
            )
        conn.commit(); conn.close()
    except Exception as e: print("hist:", e)

def tc(s):
    if not s: return s
    pre = {"a","as","o","os","da","das","do","dos","de","em","no","na","nos","nas",
           "ao","aos","por","e","ou","com","sem","sob","sobre","entre","para",
           "pelo","pela","pelos","pelas","num","numa"}
    return " ".join(w.capitalize() if i==0 or w.lower() not in pre else w.lower()
                    for i, w in enumerate(s.strip().split()))

# ── Helpers UI ────────────────────────────────────────────────────────────────
def lbl(t, color=TEXT2, size=12, bold=True):
    return ft.Text(t, size=size, color=color,
                   weight=ft.FontWeight.W_600 if bold else ft.FontWeight.NORMAL)

def inp(hint="", value="", pw=False, expand=True, width=None, ro=False, on_submit=None):
    tf = ft.TextField(
        hint_text=hint, value=value, password=pw, can_reveal_password=pw,
        bgcolor=BG2, color=TEXT, border_color=BORDER, focused_border_color=ACCENT,
        border_radius=8, content_padding=ft.padding.symmetric(horizontal=12,vertical=10),
        text_size=14, expand=expand, width=width, read_only=ro,
        hint_style=ft.TextStyle(color=TEXT3),
    )
    if on_submit: tf.on_submit = on_submit
    return tf

def dd(options, value="", hint="— Selecione —", expand=True, width=None, on_change=None):
    return ft.Dropdown(
        options=[ft.dropdown.Option(str(o)) for o in options],
        value=value or None, hint_text=hint,
        bgcolor=BG2, color=TEXT, border_color=BORDER, focused_border_color=ACCENT,
        border_radius=8, content_padding=ft.padding.symmetric(horizontal=12,vertical=4),
        text_size=14, expand=expand, width=width, on_change=on_change,
    )

def btn(t, on_click=None, bg=None, expand=False, icon=None, small=False):
    return ft.ElevatedButton(
        text=t, on_click=on_click, expand=expand, icon=icon,
        style=ft.ButtonStyle(
            color={ft.ControlState.DEFAULT: WHITE},
            bgcolor={ft.ControlState.DEFAULT: bg or "#1a4a8c",
                     ft.ControlState.HOVERED: "#2255aa"},
            shape=ft.RoundedRectangleBorder(radius=8),
            padding=ft.padding.symmetric(horizontal=10 if small else 16,
                                          vertical=6  if small else 10),
        ),
    )

def sit_badge(s):
    u = (s or "").upper()
    bg, fg = SIT_COR.get(u, (CARD, TEXT2))
    return ft.Container(
        ft.Text(s or "—", size=11, color=fg, weight=ft.FontWeight.W_700),
        bgcolor=bg, border_radius=12,
        padding=ft.padding.symmetric(horizontal=8, vertical=2),
    )

def card_section(titulo, conteudo):
    return ft.Container(
        ft.Column([
            ft.Row([ft.Icon(ft.Icons.LABEL, color=ACCENT, size=14),
                    ft.Text(titulo, size=13, color=ACCENT, weight=ft.FontWeight.W_700)]),
            ft.Divider(height=1, color=BORDER),
            *conteudo,
        ], spacing=8),
        bgcolor=BG3, border_radius=10,
        border=ft.border.all(1, BORDER),
        padding=12, margin=ft.margin.only(bottom=10),
    )

def row_2(a, b): return ft.Row([a, b], spacing=10)

# ── APP ───────────────────────────────────────────────────────────────────────
class App:
    def __init__(self, page: ft.Page):
        self.page = page
        self.usuario = None
        self.login_str = ""
        load_config()
        page.title = "Gerenciamento dos Reparos"
        page.theme_mode = ft.ThemeMode.DARK
        page.bgcolor = BG
        page.padding = 0
        page.theme = ft.Theme(color_scheme_seed=ACCENT)
        if has_config():
            init_db()
            self._ir_login()
        else:
            self._ir_config(primeiro_acesso=True)

    # ── SNACKBAR ──────────────────────────────────────────────────────────────
    def snack(self, msg, cor="#2e7d32"):
        self.page.snack_bar = ft.SnackBar(ft.Text(msg, color=WHITE), bgcolor=cor, duration=3000)
        self.page.snack_bar.open = True
        self.page.update()

    def clear(self):
        self.page.controls.clear()
        self.page.appbar = None
        self.page.floating_action_button = None

    # ── CONFIG DO BANCO ───────────────────────────────────────────────────────
    def _ir_config(self, primeiro_acesso=False):
        self.clear()
        e_path = inp(
            hint="Ex: C:\\Users\\...\\OneDrive\\SOLICITACOES.db",
            value=get_db_path() or "",
        )
        e_adm = inp("Senha do administrador", pw=True)
        lbl_err = ft.Text("", color="#ff6666", size=13)
        lbl_status = ft.Text("", color="#44cc88", size=12)

        def salvar(e):
            if not primeiro_acesso and (e_adm.value or "") != ADMIN_PASSWORD:
                lbl_err.value = "❌  Senha de administrador incorreta."
                self.page.update(); return
            p = (e_path.value or "").strip()
            if not p:
                lbl_err.value = "⚠  Informe o caminho do banco."
                self.page.update(); return
            save_config(p)
            ok = init_db()
            if ok:
                lbl_err.value = ""
                lbl_status.value = f"✅  Banco configurado: {Path(p).name}"
                self.page.update()
                import time; time.sleep(1)
                self._ir_login()
            else:
                lbl_err.value = "❌  Não foi possível criar/abrir o banco neste caminho."
                self.page.update()

        titulo = "Primeira Configuração" if primeiro_acesso else "Configurar Banco"
        self.page.appbar = ft.AppBar(
            title=ft.Text(titulo, color=TEXT), bgcolor=CARD,
            leading=None if primeiro_acesso else ft.IconButton(
                ft.Icons.ARROW_BACK, icon_color=TEXT, on_click=lambda e: self._ir_login()
            ),
        )
        conteudo = [
            ft.Text("Informe o caminho completo do arquivo .db", size=12, color=TEXT3),
            ft.Text("(pode ser na pasta do OneDrive ou qualquer local)", size=11, color=TEXT3),
            ft.Container(height=4),
            lbl("Caminho do Banco (.db) *"), e_path,
        ]
        if not primeiro_acesso:
            conteudo += [lbl("Senha do Administrador *"), e_adm]
        conteudo += [
            lbl_err, lbl_status,
            btn("💾  Salvar e Conectar", on_click=salvar, bg=SUCCESS, expand=True),
            ft.Container(height=8),
            ft.Container(
                ft.Column([
                    ft.Text("💡  Dicas:", size=12, color=ACCENT, weight=ft.FontWeight.W_600),
                    ft.Text("• Windows: C:\\Users\\nome\\OneDrive\\pasta\\banco.db", size=11, color=TEXT3),
                    ft.Text("• OneDrive sincroniza automaticamente com o PC", size=11, color=TEXT3),
                    ft.Text("• O banco já existente do desktop também funciona", size=11, color=TEXT3),
                ], spacing=4),
                bgcolor=BG2, border_radius=8, padding=12,
                border=ft.border.all(1, BORDER),
            ),
        ]
        self.page.controls.append(ft.ListView([
            ft.Container(ft.Column(conteudo, spacing=8), padding=16)
        ], expand=True, padding=0))
        self.page.update()

    # ── LOGIN ─────────────────────────────────────────────────────────────────
    def _ir_login(self):
        self.clear()
        e_login = inp("Login")
        e_senha = inp("Senha", pw=True)
        lbl_err = ft.Text("", color="#ff6666", size=13)

        def entrar(e=None):
            l = (e_login.value or "").strip()
            s = (e_senha.value or "")
            if not l or not s:
                lbl_err.value = "⚠  Preencha login e senha."; self.page.update(); return
            if not has_config():
                self._ir_config(primeiro_acesso=True); return
            u = autenticar(l, s)
            if u:
                self.usuario = u; self.login_str = l
                self._ir_lista()
            else:
                lbl_err.value = "❌  Login ou senha inválidos."; self.page.update()

        e_senha.on_submit = entrar

        db_status = ft.Container(
            ft.Row([
                ft.Icon(ft.Icons.CHECK_CIRCLE if has_config() else ft.Icons.WARNING,
                        color="#44cc88" if has_config() else "#ffaa44", size=16),
                ft.Text(
                    f"Banco: {Path(get_db_path()).name}" if has_config() else "Banco não configurado",
                    size=11, color="#44cc88" if has_config() else "#ffaa44",
                ),
            ], spacing=6),
            bgcolor=BG2, border_radius=6, padding=ft.padding.symmetric(horizontal=10, vertical=6),
            border=ft.border.all(1, BORDER),
        )

        self.page.controls.append(ft.Column([
            ft.Container(
                ft.Column([
                    ft.Icon(ft.Icons.LIST_ALT, color=ACCENT, size=52),
                    ft.Text("Gerenciamento dos Reparos", size=18, color=ACCENT,
                            weight=ft.FontWeight.BOLD, text_align=ft.TextAlign.CENTER),
                    ft.Text("Prefeitura de Maricá", size=12, color=TEXT3,
                            text_align=ft.TextAlign.CENTER),
                ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=6),
                bgcolor=CARD, padding=24,
                border=ft.border.only(bottom=ft.BorderSide(1, BORDER)),
            ),
            ft.Container(ft.Column([
                db_status,
                ft.Container(height=8),
                lbl("Login"), e_login,
                lbl("Senha"), e_senha,
                lbl_err,
                btn("▶  ENTRAR", on_click=entrar, bg=SUCCESS, expand=True),
                ft.Divider(color=BORDER),
                ft.Row([
                    ft.TextButton("➕  Cadastrar Usuário",
                        on_click=lambda e: self._ir_cadastro(),
                        style=ft.ButtonStyle(color=TEXT2)),
                    ft.TextButton("⚙️  Configurar Banco",
                        on_click=lambda e: self._ir_config(),
                        style=ft.ButtonStyle(color=TEXT3)),
                ], alignment=ft.MainAxisAlignment.SPACE_AROUND),
            ], spacing=8), padding=20, expand=True),
        ], spacing=0, expand=True, scroll=ft.ScrollMode.AUTO))
        self.page.update()

    # ── CADASTRO ──────────────────────────────────────────────────────────────
    def _ir_cadastro(self):
        self.clear()
        e_adm  = inp("Senha do administrador", pw=True)
        e_nome = inp("Nome"); e_sob = inp("Sobrenome")
        e_log  = inp("Login")
        e_sen  = inp("Senha (mín. 4 caracteres)", pw=True)
        e_con  = inp("Confirmar senha", pw=True)
        lbl_er = ft.Text("", color="#ff6666", size=13)

        def salvar(e):
            if (e_adm.value or "") != ADMIN_PASSWORD:
                lbl_er.value = "❌  Senha de administrador incorreta."; self.page.update(); return
            n,s,l,p,c = (e_nome.value or "").strip(),(e_sob.value or "").strip(),\
                         (e_log.value or "").strip(),(e_sen.value or ""),(e_con.value or "")
            if not all([n,s,l,p,c]):
                lbl_er.value = "⚠  Preencha todos os campos."; self.page.update(); return
            if len(p) < 4:
                lbl_er.value = "⚠  Senha mínimo 4 caracteres."; self.page.update(); return
            if p != c:
                lbl_er.value = "❌  Senhas não coincidem."; self.page.update(); return
            ok, msg = cadastrar_usuario(n, s, l, p)
            if ok: self.snack(f"✅  Usuário '{n} {s}' cadastrado!"); self._ir_login()
            else:  lbl_er.value = f"❌  {msg}"; self.page.update()

        self.page.appbar = ft.AppBar(
            title=ft.Text("Cadastrar Usuário", color=TEXT), bgcolor=CARD,
            leading=ft.IconButton(ft.Icons.ARROW_BACK, icon_color=TEXT,
                                  on_click=lambda e: self._ir_login()),
        )
        self.page.controls.append(ft.ListView([ft.Container(ft.Column([
            lbl("Senha do Administrador *"), e_adm,
            row_2(ft.Column([lbl("Nome *"), e_nome], expand=True),
                  ft.Column([lbl("Sobrenome *"), e_sob], expand=True)),
            lbl("Login *"), e_log, lbl("Senha *"), e_sen,
            lbl("Confirmar Senha *"), e_con, lbl_er,
            btn("💾  Salvar Usuário", on_click=salvar, bg=SUCCESS, expand=True),
            ft.Container(height=20),
        ], spacing=8), padding=16)], expand=True, padding=0))
        self.page.update()

    # ── LISTA ─────────────────────────────────────────────────────────────────
    def _ir_lista(self):
        self.clear()
        self._pg = 1; self._per = 30
        self._f = {"q":"","nucleo":"","situacao":"","categoria":"","bairro":""}

        self._tf_q = ft.TextField(
            hint_text="🔍  Nº, solicitante, endereço...",
            bgcolor=BG2, color=TEXT, border_color=BORDER, focused_border_color=ACCENT,
            border_radius=8, content_padding=ft.padding.symmetric(horizontal=12,vertical=8),
            text_size=13, on_change=self._on_q,
        )
        self._dd_nuc = dd([""] + NUCLEOS, hint="Núcleo", on_change=self._on_f)
        self._dd_sit = dd([""] + SITUACOES, hint="Situação", on_change=self._on_f)
        self._dd_cat = dd([""] + list(CAT_SUBCAT.keys()), hint="Categoria", on_change=self._on_f)
        self._tf_bai = ft.TextField(
            hint_text="Bairro", bgcolor=BG2, color=TEXT, border_color=BORDER,
            focused_border_color=ACCENT, border_radius=8,
            content_padding=ft.padding.symmetric(horizontal=10,vertical=8),
            text_size=13, expand=True, on_change=self._on_q,
        )
        self._lbl_tot = ft.Text("", size=12, color=TEXT3)
        self._col_lista = ft.Column([], spacing=6, scroll=ft.ScrollMode.AUTO, expand=True)
        self._mostrar_filtros = False
        self._filtros_ext = ft.Container(
            ft.Column([
                row_2(self._dd_cat, self._tf_bai),
                row_2(self._dd_nuc, self._dd_sit),
            ], spacing=8),
            padding=ft.padding.only(left=12,right=12,bottom=10),
            bgcolor=BG2, visible=False,
        )

        self.page.appbar = ft.AppBar(
            title=ft.Text("Solicitações", color=TEXT), bgcolor=CARD,
            actions=[
                ft.IconButton(ft.Icons.ADD_CIRCLE_OUTLINE, icon_color="#4caf50",
                              tooltip="Nova", on_click=lambda e: self._ir_form("novo")),
                ft.IconButton(ft.Icons.TABLE_CHART, icon_color="#4caf50",
                              tooltip="Exportar Excel", on_click=lambda e: self._exportar_excel()),
                ft.IconButton(ft.Icons.SETTINGS, icon_color=TEXT3,
                              tooltip="Configurações", on_click=lambda e: self._ir_config()),
                ft.IconButton(ft.Icons.LOGOUT, icon_color=TEXT3,
                              tooltip="Sair", on_click=lambda e: self._ir_login()),
            ],
        )
        self.page.controls.append(ft.Column([
            ft.Container(ft.Column([
                ft.Row([
                    ft.Container(ft.Row([
                        ft.Icon(ft.Icons.SEARCH, color=TEXT3, size=18),
                        ft.Container(self._tf_q, expand=True),
                    ], spacing=8), expand=True,
                    bgcolor=BG2, border_radius=8,
                    border=ft.border.all(1, BORDER),
                    padding=ft.padding.only(left=10,right=4)),
                    ft.IconButton(ft.Icons.FILTER_LIST,
                                  icon_color=ACCENT, icon_size=22,
                                  on_click=self._toggle_filtros,
                                  tooltip="Filtros"),
                ], spacing=6),
                self._filtros_ext,
                self._lbl_tot,
            ], spacing=6),
            bgcolor=BG2, padding=ft.padding.symmetric(horizontal=12,vertical=10),
            border=ft.border.only(bottom=ft.BorderSide(1,BORDER))),
            ft.Container(self._col_lista, expand=True, padding=8),
        ], spacing=0, expand=True))
        self.page.update()
        self._carregar_lista()

    def _toggle_filtros(self, e):
        self._mostrar_filtros = not self._mostrar_filtros
        self._filtros_ext.visible = self._mostrar_filtros
        self.page.update()

    def _on_q(self, e):
        self._f["q"] = (self._tf_q.value or "").strip()
        self._f["bairro"] = (self._tf_bai.value or "").strip()
        self._pg = 1; self._carregar_lista()

    def _on_f(self, e):
        self._f["nucleo"]   = self._dd_nuc.value or ""
        self._f["situacao"] = self._dd_sit.value or ""
        self._f["categoria"]= self._dd_cat.value or ""
        self._pg = 1; self._carregar_lista()

    def _carregar_lista(self):
        q=self._f["q"]; nuc=self._f["nucleo"]; sit=self._f["situacao"]
        cat=self._f["categoria"]; bai=self._f["bairro"]
        offset=(self._pg-1)*self._per
        w,p=["1=1"],[]
        if q:
            w.append("(solicitacao LIKE ? OR solicitante LIKE ? OR endereco LIKE ? OR observacao LIKE ?)")
            p+=[f"%{q}%"]*4
        if nuc: w.append("nucleo=?"); p.append(nuc)
        if sit: w.append("situacao=?"); p.append(sit)
        if cat: w.append("categoria=?"); p.append(cat)
        if bai: w.append("bairro LIKE ?"); p.append(f"%{bai}%")
        sw=" AND ".join(w)
        try:
            conn=get_conn()
            total=conn.execute(f"SELECT COUNT(*) FROM solicitacoes WHERE {sw}",p).fetchone()[0]
            rows=conn.execute(
                f"SELECT id,data_abertura,solicitacao,nucleo,solicitante,endereco,"
                f"bairro,situacao,categoria,encarregado,"
                f"(SELECT COUNT(*) FROM fotos f WHERE f.solicitacao_id=s.id) as qtd_fotos,"
                f"(SELECT COUNT(*) FROM materiais m WHERE m.solicitacao_id=s.id) as qtd_mat "
                f"FROM solicitacoes s WHERE {sw} "
                f"ORDER BY CAST(substr(solicitacao,1,5) AS INTEGER) DESC "
                f"LIMIT ? OFFSET ?",
                p+[self._per,offset]
            ).fetchall()
            conn.close()
        except Exception as ex:
            self._col_lista.controls=[ft.Text(f"Erro: {ex}", color="#ff6666")]
            self.page.update(); return

        self._lbl_tot.value=f"{total} registro(s)"
        self._col_lista.controls.clear()
        if not rows:
            self._col_lista.controls.append(
                ft.Container(ft.Text("Nenhum registro encontrado.",
                             color=TEXT3, text_align=ft.TextAlign.CENTER),
                             alignment=ft.alignment.center, padding=40))
        else:
            for r in rows:
                self._col_lista.controls.append(self._card_sol(dict(r)))
        total_pgs=(total+self._per-1)//self._per
        if total_pgs>1:
            self._col_lista.controls.append(
                ft.Row([
                    ft.IconButton(ft.Icons.CHEVRON_LEFT, disabled=self._pg<=1,
                                  on_click=lambda e: self._pag(-1)),
                    ft.Text(f"{self._pg}/{total_pgs}", color=TEXT2, size=13),
                    ft.IconButton(ft.Icons.CHEVRON_RIGHT, disabled=self._pg>=total_pgs,
                                  on_click=lambda e: self._pag(1)),
                ], alignment=ft.MainAxisAlignment.CENTER))
        self.page.update()

    def _pag(self, d): self._pg+=d; self._carregar_lista()

    def _card_sol(self, r):
        sit=(r.get("situacao") or "").upper()
        bg_map={"ABERTO":"#2a2000","EM EXECUÇÃO":"#0a1a2e",
                "CANCELADO":"#2a0d0d","FINALIZADO":"#0d2a18"}
        bg=bg_map.get(sit, BG3); rid=r["id"]
        info2 = []
        if r.get("encarregado"): info2.append(r["encarregado"])
        if r.get("qtd_fotos",0)>0: info2.append(f"📷{r['qtd_fotos']}")
        if r.get("qtd_mat",0)>0:   info2.append(f"🧱{r['qtd_mat']}")
        return ft.Container(
            ft.Column([
                ft.Row([
                    ft.Text(r.get("solicitacao","—"), size=14, color="#a0c8ff",
                            weight=ft.FontWeight.BOLD, expand=True),
                    sit_badge(r.get("situacao","—")),
                ]),
                ft.Row([ft.Text(r.get("data_abertura",""),size=11,color=TEXT3),
                        ft.Text("·",color=TEXT3),
                        ft.Text(r.get("nucleo","—"),size=11,color=TEXT2)], spacing=4),
                ft.Text(r.get("solicitante","—"), size=12, color=TEXT2),
                ft.Text(
                    (r.get("endereco") or "") +
                    ((", "+r["bairro"]) if r.get("bairro") else ""),
                    size=11, color=TEXT3, max_lines=1,
                    overflow=ft.TextOverflow.ELLIPSIS),
                ft.Row([
                    ft.Text(r.get("categoria","—"), size=11, color=TEXT3, expand=True),
                    ft.Text(" · ".join(info2), size=11, color=TEXT3),
                ]),
            ], spacing=3),
            bgcolor=bg, border_radius=10, border=ft.border.all(1,BORDER),
            padding=12, on_click=lambda e,i=rid: self._ir_form("editar",i),
        )

    # ── EXPORTAR EXCEL ────────────────────────────────────────────────────────
    def _exportar_excel(self):
        def _bg():
            try:
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
            except ImportError:
                self.snack("❌  openpyxl não instalado. Execute: pip install openpyxl", "#b53030")
                return
            try:
                conn = get_conn()
                rows = conn.execute(
                    "SELECT s.data_abertura,s.solicitacao,s.nucleo,s.solicitante,"
                    "s.endereco,s.bairro,s.distrito,s.categoria,s.subcategoria,"
                    "s.encarregado,s.situacao,s.data_inicio,s.data_conclusao,"
                    "s.observacao,s.localizacao,"
                    "(SELECT COUNT(*) FROM fotos f WHERE f.solicitacao_id=s.id) as qtd_fotos,"
                    "(SELECT COUNT(*) FROM materiais m WHERE m.solicitacao_id=s.id) as qtd_mat "
                    "FROM solicitacoes s ORDER BY s.solicitacao DESC"
                ).fetchall()
                mat_rows = conn.execute(
                    "SELECT s.solicitacao,s.nucleo,m.nome,m.quantidade,m.tipo "
                    "FROM materiais m JOIN solicitacoes s ON s.id=m.solicitacao_id "
                    "ORDER BY s.solicitacao,m.id"
                ).fetchall()
                conn.close()

                wb = Workbook()
                thin = Side(style="thin", color="2E2E55")
                bord = Border(left=thin, right=thin, top=thin, bottom=thin)
                ws = wb.active; ws.title = "Solicitações"
                hdrs = ["DATA ABERTURA","SOLICITAÇÃO","NÚCLEO","SOLICITANTE","ENDEREÇO",
                        "BAIRRO","DISTRITO","CATEGORIA","SUBCATEGORIA","ENCARREGADO",
                        "SITUAÇÃO","DATA INÍCIO","DATA CONCLUSÃO","OBSERVAÇÃO",
                        "LOCALIZAÇÃO","FOTOS","MAT"]
                for ci,txt in enumerate(hdrs,1):
                    c=ws.cell(row=1,column=ci,value=txt)
                    c.fill=PatternFill("solid",fgColor="0D0D3A")
                    c.font=Font(name="Segoe UI",bold=True,color="A0C4FF",size=11)
                    c.alignment=Alignment(horizontal="center",vertical="center"); c.border=bord
                ws.row_dimensions[1].height=30
                CORES={"FINALIZADO":("1A4A2E","7AFF7A"),"CANCELADO":("4A1A1A","FF7A7A"),
                       "EM EXECUÇÃO":("0D2A4A","88CCFF"),"ABERTO":("3A3000","FFDD66")}
                for ri,row in enumerate(rows,2):
                    sit=(row["situacao"] or "").upper()
                    bg,fg=CORES.get(sit,("1E1E2E","D0D8F0") if ri%2==0 else ("252535","D0D8F0"))
                    fill=PatternFill("solid",fgColor=bg); font=Font(name="Segoe UI",color=fg,size=10)
                    vals=[row["data_abertura"] or "",row["solicitacao"] or "",row["nucleo"] or "",
                          row["solicitante"] or "",row["endereco"] or "",row["bairro"] or "",
                          row["distrito"] or "",row["categoria"] or "",row["subcategoria"] or "",
                          row["encarregado"] or "",row["situacao"] or "",row["data_inicio"] or "",
                          row["data_conclusao"] or "",(row["observacao"] or "")[:200],
                          (row["localizacao"] or ""),row["qtd_fotos"] or 0,row["qtd_mat"] or 0]
                    for ci,val in enumerate(vals,1):
                        c=ws.cell(row=ri,column=ci,value=val)
                        c.fill=fill; c.font=font
                        c.alignment=Alignment(vertical="center"); c.border=bord
                    ws.row_dimensions[ri].height=20
                for i,w in enumerate([13,14,14,18,24,14,12,22,18,14,14,12,12,32,28,6,6],1):
                    ws.column_dimensions[get_column_letter(i)].width=w
                ws.freeze_panes="A2"
                ws_m=wb.create_sheet("Materiais")
                for ci,txt in enumerate(["SOLICITACAO","NUCLEO","MATERIAL","QUANTIDADE","TIPO"],1):
                    c=ws_m.cell(row=1,column=ci,value=txt)
                    c.fill=PatternFill("solid",fgColor="0D2A4A")
                    c.font=Font(name="Segoe UI",bold=True,color="88CCFF",size=11)
                for ri,mr in enumerate(mat_rows,2):
                    for ci,val in enumerate([mr["solicitacao"] or "",mr["nucleo"] or "",
                            mr["nome"] or "",mr["quantidade"] or "",mr["tipo"] or ""],1):
                        ws_m.cell(row=ri,column=ci,value=val)
                ws_m.freeze_panes="A2"
                destino = Path(get_db_path()).parent / f"solicitacoes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                wb.save(str(destino))
                self.snack(f"✅  Excel salvo: {destino.name}")
            except Exception as ex:
                self.snack(f"❌  Erro: {ex}", "#b53030")
        threading.Thread(target=_bg, daemon=True).start()
        self.snack("⏳  Gerando Excel na pasta do banco...", "#1a4a8c")

    # ── FORMULÁRIO ────────────────────────────────────────────────────────────
    def _ir_form(self, modo="novo", sol_id=None):
        dados = {}
        if modo=="editar" and sol_id:
            conn=get_conn()
            row=conn.execute("SELECT * FROM solicitacoes WHERE id=?",(sol_id,)).fetchone()
            conn.close()
            if row: dados=dict(row)

        self.clear()
        titulo="Nova Solicitação" if modo=="novo" else f"Editar — {dados.get('solicitacao','')}"

        e_data=inp("DD/MM/AAAA", value=dados.get("data_abertura",""))
        e_sol =inp("00001/2026", value=dados.get("solicitacao",""), ro=(modo=="editar"))
        e_sol2=inp("Solicitante", value=dados.get("solicitante",""))
        e_end =inp("Endereço / Logradouro", value=dados.get("endereco",""))
        e_loc =inp("Ex: 22°57'43.6\"S 42°57'56.6\"W", value=dados.get("localizacao",""))
        e_obs =ft.TextField(
            hint_text="Observações...", value=dados.get("observacao",""),
            bgcolor=BG2, color=TEXT, border_color=BORDER, focused_border_color=ACCENT,
            border_radius=8, multiline=True, min_lines=3, max_lines=6,
            content_padding=ft.padding.symmetric(horizontal=12,vertical=10), text_size=14,
        )
        e_di=inp("DD/MM/AAAA", value=dados.get("data_inicio",""))
        e_dc=inp("DD/MM/AAAA", value=dados.get("data_conclusao",""))
        d_nuc=dd(NUCLEOS, value=dados.get("nucleo",""))
        d_dis=inp("Distrito", value=dados.get("distrito",""), ro=True)
        d_bai=dd(DISTRITO_BAIRROS.get(dados.get("nucleo",""),[]),
                 value=dados.get("bairro",""))
        d_cat=dd(list(CAT_SUBCAT.keys()), value=dados.get("categoria",""))
        d_sub=dd(CAT_SUBCAT.get(dados.get("categoria",""),[]),
                 value=dados.get("subcategoria",""))
        d_sit=dd(SITUACOES, value=dados.get("situacao",""))
        e_enc=inp("Encarregado", value=dados.get("encarregado",""))
        lbl_er=ft.Text("", color="#ff6666", size=13)

        def on_nuc(e):
            nuc=d_nuc.value or ""
            d_dis.value=NUCLEO_DISTRITO.get(nuc,"")
            bairros=DISTRITO_BAIRROS.get(nuc,[])
            d_bai.options=[ft.dropdown.Option(b) for b in bairros]
            d_bai.value=None; self.page.update()

        def on_cat(e):
            subs=CAT_SUBCAT.get(d_cat.value or "",[])
            d_sub.options=[ft.dropdown.Option(s) for s in subs]
            d_sub.value=None; self.page.update()

        d_nuc.on_change=on_nuc; d_cat.on_change=on_cat

        if modo=="novo":
            e_data.value=datetime.now().strftime("%d/%m/%Y")
            e_sol.value=gerar_proxima()
            self.page.update()

        def _snap():
            return {"Data Abertura":e_data.value or "","Solicitação":e_sol.value or "",
                    "Núcleo":d_nuc.value or "","Solicitante":e_sol2.value or "",
                    "Endereço":e_end.value or "","Bairro":d_bai.value or "",
                    "Distrito":d_dis.value or "","Categoria":d_cat.value or "",
                    "Subcategoria":d_sub.value or "","Encarregado":e_enc.value or "",
                    "Situação":d_sit.value or "","Data Início":e_di.value or "",
                    "Data Conclusão":e_dc.value or "","Observação":e_obs.value or "",
                    "Localização":e_loc.value or ""}

        _orig = _snap() if modo=="editar" else {}

        def _salvar(e=None):
            da=(e_data.value or "").strip(); s=(e_sol.value or "").strip()
            if not da or not s:
                lbl_er.value="⚠  Data e Número são obrigatórios."; self.page.update(); return
            sit=(d_sit.value or "").upper()
            if "EXECU" in sit and not (e_di.value or "").strip():
                lbl_er.value="⚠  Em Execução requer Data de Início."; self.page.update(); return
            if "FINALIZ" in sit and not (e_dc.value or "").strip():
                lbl_er.value="⚠  Finalizado requer Data de Conclusão."; self.page.update(); return
            reg={"data_abertura":da,"solicitacao":s,
                 "nucleo":d_nuc.value or "","solicitante":tc(e_sol2.value or ""),
                 "endereco":tc(e_end.value or ""),"bairro":d_bai.value or "",
                 "distrito":d_dis.value or "","categoria":d_cat.value or "",
                 "subcategoria":d_sub.value or "","encarregado":tc(e_enc.value or ""),
                 "situacao":(d_sit.value or "").upper(),
                 "data_inicio":(e_di.value or "").strip(),
                 "data_conclusao":(e_dc.value or "").strip(),
                 "observacao":tc(e_obs.value or ""),
                 "localizacao":(e_loc.value or "").strip()}
            try:
                conn=get_conn()
                if modo=="novo":
                    cur=conn.execute("""INSERT INTO solicitacoes
                        (data_abertura,solicitacao,nucleo,solicitante,endereco,bairro,
                         distrito,categoria,subcategoria,encarregado,situacao,
                         data_inicio,data_conclusao,observacao,localizacao)
                        VALUES (:data_abertura,:solicitacao,:nucleo,:solicitante,:endereco,
                         :bairro,:distrito,:categoria,:subcategoria,:encarregado,:situacao,
                         :data_inicio,:data_conclusao,:observacao,:localizacao)""", reg)
                    conn.commit(); new_id=cur.lastrowid; conn.close()
                    self.snack(f"✅  Solicitação {s} criada!")
                    self._ir_form("editar", new_id)
                else:
                    reg["id"]=sol_id
                    conn.execute("""UPDATE solicitacoes SET
                        data_abertura=:data_abertura,nucleo=:nucleo,
                        solicitante=:solicitante,endereco=:endereco,bairro=:bairro,
                        distrito=:distrito,categoria=:categoria,subcategoria=:subcategoria,
                        encarregado=:encarregado,situacao=:situacao,
                        data_inicio=:data_inicio,data_conclusao=:data_conclusao,
                        observacao=:observacao,localizacao=:localizacao,
                        atualizado_em=datetime('now','localtime') WHERE id=:id""", reg)
                    conn.commit()
                    novo=_snap()
                    diff=[(k,_orig.get(k,""),novo.get(k,""))
                          for k in novo if _orig.get(k,"")!=novo.get(k,"")]
                    conn.close()
                    salvar_historico(sol_id,s,self.login_str,diff)
                    self.snack(f"✅  Solicitação {s} atualizada!")
                    self._ir_lista()
            except sqlite3.IntegrityError:
                lbl_er.value=f"❌  Solicitação '{s}' já existe!"; self.page.update()
            except Exception as ex:
                lbl_er.value=f"❌  Erro: {ex}"; self.page.update()

        def _excluir_dlg():
            def _ok(ev):
                dlg.open=False; self.page.update()
                try:
                    conn=get_conn()
                    row=conn.execute("SELECT * FROM solicitacoes WHERE id=?",(sol_id,)).fetchone()
                    if row:
                        conn.execute(
                            "INSERT INTO excluidos (solicitacao,excluido_por,excluido_em,dados_json)"
                            " VALUES (?,?,?,?)",
                            (row["solicitacao"],self.login_str,
                             datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                             json.dumps(dict(row),ensure_ascii=False)))
                        conn.execute("DELETE FROM solicitacoes WHERE id=?",(sol_id,))
                        conn.commit()
                    conn.close()
                    self.snack("🗑️  Excluído."); self._ir_lista()
                except Exception as ex:
                    self.snack(f"❌  {ex}","#b53030")
            dlg=ft.AlertDialog(
                title=ft.Text("Confirmar Exclusão",color=TEXT),
                content=ft.Text(f"Excluir {dados.get('solicitacao','')}?",color=TEXT2),
                bgcolor=CARD,
                actions=[
                    ft.TextButton("Cancelar",
                        on_click=lambda ev:(setattr(dlg,"open",False),self.page.update())),
                    ft.ElevatedButton("Excluir",on_click=_ok,
                        style=ft.ButtonStyle(
                            bgcolor={ft.ControlState.DEFAULT:DANGER},
                            color={ft.ControlState.DEFAULT:WHITE})),
                ],
                actions_alignment=ft.MainAxisAlignment.END,
            )
            self.page.overlay.append(dlg); dlg.open=True; self.page.update()

        # Botão fotos e materiais (só edição)
        extra = []
        if modo=="editar" and sol_id:
            extra = [
                ft.Row([
                    btn("📷  Fotos", on_click=lambda e: self._ir_fotos(sol_id, dados.get("solicitacao","")),
                        bg="#1a3a6b", expand=True),
                    btn("🧱  Materiais", on_click=lambda e: self._ir_materiais(sol_id, dados.get("solicitacao","")),
                        bg="#1a3a6b", expand=True),
                ], spacing=8),
                btn("📋  Histórico", on_click=lambda e: self._ir_historico(sol_id),
                    bg="#2a1a4a", expand=True),
            ]

        actions = [ft.IconButton(ft.Icons.SAVE, icon_color="#4caf50",
                                  tooltip="Salvar", on_click=_salvar)]
        if modo=="editar":
            actions.append(ft.IconButton(ft.Icons.DELETE_OUTLINE, icon_color="#ef5350",
                                          tooltip="Excluir", on_click=lambda e: _excluir_dlg()))

        self.page.appbar = ft.AppBar(
            title=ft.Text(titulo, color=TEXT, size=14), bgcolor=CARD,
            leading=ft.IconButton(ft.Icons.ARROW_BACK, icon_color=TEXT,
                                  on_click=lambda e: self._ir_lista()),
            actions=actions,
        )
        self.page.controls.append(ft.ListView([ft.Container(ft.Column([
            card_section("🪪  Identificação", [
                row_2(ft.Column([lbl("Data Abertura *"),e_data],expand=True),
                      ft.Column([lbl("Nº Solicitação *"),e_sol],expand=True)),
                lbl("Solicitante"), e_sol2,
            ]),
            card_section("📍  Localização", [
                lbl("Endereço"), e_end,
                row_2(ft.Column([lbl("Núcleo"),d_nuc],expand=True),
                      ft.Column([lbl("Distrito"),d_dis],expand=True)),
                lbl("Bairro"), d_bai,
                lbl("Coordenadas"),
                ft.Row([
                    ft.Container(e_loc, expand=True),
                    ft.IconButton(ft.Icons.MAP, icon_color=ACCENT, tooltip="Google Maps",
                        on_click=lambda e: self._abrir_maps(
                            e_loc.value or f"{e_end.value} {d_bai.value}")),
                ], spacing=4),
            ]),
            card_section("🔧  Serviço", [
                row_2(ft.Column([lbl("Categoria"),d_cat],expand=True),
                      ft.Column([lbl("Subcategoria"),d_sub],expand=True)),
                lbl("Encarregado"), e_enc,
            ]),
            card_section("📊  Status", [
                lbl("Situação"), d_sit,
                row_2(ft.Column([lbl("Data Início"),e_di],expand=True),
                      ft.Column([lbl("Data Conclusão"),e_dc],expand=True)),
                lbl("Observação"), e_obs,
            ]),
            *extra,
            lbl_er,
            btn("💾  Salvar", on_click=_salvar, bg=SUCCESS, expand=True),
            ft.Container(height=24),
        ], spacing=0), padding=12)], expand=True, padding=0))
        self.page.update()

    def _abrir_maps(self, loc):
        import urllib.parse, webbrowser
        if loc and loc.strip():
            webbrowser.open(f"https://www.google.com/maps/search/?api=1&query={urllib.parse.quote(loc.strip())}")

    # ── FOTOS ─────────────────────────────────────────────────────────────────
    def _ir_fotos(self, sol_id, sol_num):
        self.clear()
        foto_dir = Path(get_foto_dir())
        foto_dir.mkdir(parents=True, exist_ok=True)
        col_fotos = ft.Column([], spacing=8, scroll=ft.ScrollMode.AUTO)
        lbl_er = ft.Text("", color="#ff6666", size=13)

        def _carregar():
            col_fotos.controls.clear()
            try:
                conn=get_conn()
                fotos=conn.execute(
                    "SELECT id,nome_arquivo,descricao FROM fotos WHERE solicitacao_id=? ORDER BY id",
                    (sol_id,)
                ).fetchall()
                conn.close()
            except: fotos=[]
            if not fotos:
                col_fotos.controls.append(
                    ft.Text("Nenhuma foto adicionada ainda.",color=TEXT3,size=13))
            for f in fotos:
                fpath=foto_dir/f["nome_arquivo"]
                row=ft.Container(ft.Row([
                    ft.Icon(ft.Icons.IMAGE, color=ACCENT, size=32) if not fpath.exists()
                        else ft.Image(src=str(fpath), width=60, height=60, fit=ft.ImageFit.COVER,
                                      border_radius=6),
                    ft.Column([
                        ft.Text(f["nome_arquivo"],size=12,color=TEXT2,
                                overflow=ft.TextOverflow.ELLIPSIS),
                        ft.Text(f["descricao"] or "",size=11,color=TEXT3),
                    ], expand=True, spacing=2),
                    ft.IconButton(ft.Icons.DELETE, icon_color="#ef5350", icon_size=20,
                        on_click=lambda e,fid=f["id"],fn=f["nome_arquivo"]:
                            _deletar(fid,fn)),
                ], spacing=8, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                bgcolor=BG2, border_radius=8, padding=8,
                border=ft.border.all(1,BORDER))
                col_fotos.controls.append(row)
            self.page.update()

        def _deletar(fid, nome):
            try:
                conn=get_conn()
                conn.execute("DELETE FROM fotos WHERE id=?",(fid,))
                conn.commit(); conn.close()
                fp=foto_dir/nome
                if fp.exists(): fp.unlink()
                _carregar()
            except Exception as ex:
                lbl_er.value=f"❌  {ex}"; self.page.update()

        def _adicionar(e: ft.FilePickerResultEvent):
            if not e.files: return
            try:
                conn=get_conn()
                for f in e.files:
                    ext=Path(f.name).suffix.lower() or ".jpg"
                    seq=conn.execute(
                        "SELECT COUNT(*) FROM fotos WHERE solicitacao_id=?",(sol_id,)
                    ).fetchone()[0]+1
                    prefixo=sol_num.replace("/","-")
                    nome=f"{prefixo}_{seq:03d}{ext}"
                    dest=foto_dir/nome
                    shutil.copy2(f.path, str(dest))
                    conn.execute(
                        "INSERT INTO fotos (solicitacao_id,nome_arquivo,descricao) VALUES (?,?,?)",
                        (sol_id,nome,"")
                    )
                conn.commit(); conn.close()
                _carregar()
                self.snack(f"✅  {len(e.files)} foto(s) adicionada(s)!")
            except Exception as ex:
                lbl_er.value=f"❌  {ex}"; self.page.update()

        picker=ft.FilePicker(on_result=_adicionar)
        self.page.overlay.append(picker)

        self.page.appbar=ft.AppBar(
            title=ft.Text(f"Fotos — {sol_num}",color=TEXT,size=14), bgcolor=CARD,
            leading=ft.IconButton(ft.Icons.ARROW_BACK,icon_color=TEXT,
                                  on_click=lambda e: self._ir_form("editar",sol_id)),
            actions=[ft.IconButton(ft.Icons.ADD_PHOTO_ALTERNATE, icon_color="#4caf50",
                tooltip="Adicionar foto",
                on_click=lambda e: picker.pick_files(
                    allow_multiple=True,
                    allowed_extensions=["jpg","jpeg","png","gif","bmp","webp"]))],
        )
        self.page.controls.append(ft.Column([
            ft.Container(ft.Column([
                ft.Text(f"Fotos da solicitação {sol_num}",size=12,color=TEXT3),
                lbl_er,
            ],spacing=4), padding=ft.padding.symmetric(horizontal=14,vertical=10)),
            ft.Container(col_fotos, expand=True, padding=ft.padding.symmetric(horizontal=12)),
        ], expand=True))
        self.page.update()
        _carregar()

    # ── MATERIAIS ─────────────────────────────────────────────────────────────
    def _ir_materiais(self, sol_id, sol_num):
        self.clear()
        linhas=[]; col_mat=ft.Column([],spacing=6,scroll=ft.ScrollMode.AUTO)
        lbl_er=ft.Text("",color="#ff6666",size=13)

        def _add(nome="",qtd="",tipo=""):
            en=ft.TextField(value=nome,hint_text="Material",bgcolor=BG2,color=TEXT,
                border_color=BORDER,focused_border_color=ACCENT,border_radius=6,
                content_padding=ft.padding.symmetric(horizontal=8,vertical=8),
                text_size=13,expand=True)
            eq=ft.TextField(value=qtd,hint_text="Qtd",bgcolor=BG2,color=TEXT,
                border_color=BORDER,focused_border_color=ACCENT,border_radius=6,
                content_padding=ft.padding.symmetric(horizontal=8,vertical=8),
                text_size=13,width=80)
            et=ft.TextField(value=tipo,hint_text="Un",bgcolor=BG2,color=TEXT,
                border_color=BORDER,focused_border_color=ACCENT,border_radius=6,
                content_padding=ft.padding.symmetric(horizontal=8,vertical=8),
                text_size=13,width=80)
            d={"n":en,"q":eq,"t":et}; linhas.append(d)
            row_ref=ft.Ref()
            def rm(ev,_d=d):
                linhas.remove(_d)
                row_ref.current.visible=False; self.page.update()
            row=ft.Row([en,eq,et,ft.IconButton(ft.Icons.DELETE,icon_color="#ef5350",
                       icon_size=20,on_click=rm)],
                       spacing=4,vertical_alignment=ft.CrossAxisAlignment.CENTER,ref=row_ref)
            col_mat.controls.append(row); self.page.update()

        try:
            conn=get_conn()
            existentes=conn.execute(
                "SELECT nome,quantidade,tipo FROM materiais WHERE solicitacao_id=? ORDER BY id",
                (sol_id,)
            ).fetchall()
            conn.close()
        except: existentes=[]
        for r in existentes: _add(r["nome"],r["quantidade"] or "",r["tipo"] or "")
        if not existentes: _add()

        def _salvar(e):
            itens,erros=[],[]
            for i,d in enumerate(linhas):
                n=(d["n"].value or "").strip(); q=(d["q"].value or "").strip()
                t=(d["t"].value or "").strip()
                if not n: continue
                if not q: erros.append(f"Linha {i+1}: '{n}' sem quantidade.")
                else: itens.append((n,q,t))
            if erros: lbl_er.value="\n".join(erros); self.page.update(); return
            try:
                conn=get_conn()
                conn.execute("DELETE FROM materiais WHERE solicitacao_id=?",(sol_id,))
                for n,q,t in itens:
                    conn.execute(
                        "INSERT INTO materiais (solicitacao_id,nome,quantidade,tipo) VALUES (?,?,?,?)",
                        (sol_id,n,q,t))
                conn.commit(); conn.close()
                self.snack(f"✅  {len(itens)} material(is) salvo(s)!")
                self._ir_form("editar",sol_id)
            except Exception as ex:
                lbl_er.value=f"❌  {ex}"; self.page.update()

        self.page.appbar=ft.AppBar(
            title=ft.Text(f"Materiais — {sol_num}",color=TEXT,size=14), bgcolor=CARD,
            leading=ft.IconButton(ft.Icons.ARROW_BACK,icon_color=TEXT,
                                  on_click=lambda e: self._ir_form("editar",sol_id)),
            actions=[
                ft.IconButton(ft.Icons.ADD,icon_color="#4caf50",on_click=lambda e: _add()),
                ft.IconButton(ft.Icons.SAVE,icon_color=ACCENT,on_click=_salvar),
            ],
        )
        self.page.controls.append(ft.ListView([ft.Container(ft.Column([
            ft.Text("Materiais usados na solicitação",size=12,color=TEXT3),
            ft.Container(height=4),col_mat,lbl_er,
            ft.Container(height=8),
            btn("💾  Salvar Materiais",on_click=_salvar,bg=SUCCESS,expand=True),
            ft.Container(height=24),
        ],spacing=8),padding=14)],expand=True,padding=0))
        self.page.update()

    # ── HISTÓRICO ─────────────────────────────────────────────────────────────
    def _ir_historico(self, sol_id):
        self.clear()
        try:
            conn=get_conn()
            sol=conn.execute("SELECT solicitacao FROM solicitacoes WHERE id=?",(sol_id,)).fetchone()
            hist=conn.execute(
                "SELECT usuario,data_hora,campo,valor_anterior,valor_novo "
                "FROM historico WHERE solicitacao_id=? ORDER BY id DESC",
                (sol_id,)
            ).fetchall()
            conn.close()
        except: sol=None; hist=[]

        sol_num=sol["solicitacao"] if sol else str(sol_id)
        self.page.appbar=ft.AppBar(
            title=ft.Text(f"Histórico — {sol_num}",color=TEXT,size=14),bgcolor=CARD,
            leading=ft.IconButton(ft.Icons.ARROW_BACK,icon_color=TEXT,
                                  on_click=lambda e: self._ir_form("editar",sol_id)),
        )
        itens=[]
        if not hist:
            itens.append(ft.Text("Nenhuma alteração registrada.",color=TEXT3,size=13))
        for h in hist:
            itens.append(ft.Container(ft.Column([
                ft.Row([ft.Text(h["usuario"],size=11,color=ACCENT),
                        ft.Text(h["data_hora"],size=10,color=TEXT3)],spacing=6),
                ft.Text(h["campo"],size=12,color=TEXT2,weight=ft.FontWeight.W_600),
                ft.Row([
                    ft.Text(h["valor_anterior"] or "—",size=11,color="#ff8888"),
                    ft.Icon(ft.Icons.ARROW_FORWARD,size=12,color=TEXT3),
                    ft.Text(h["valor_novo"] or "—",size=11,color="#7aff9a"),
                ],spacing=4,wrap=True),
            ],spacing=3),
            bgcolor=BG2,border_radius=6,padding=10,
            border=ft.border.all(1,BORDER)))
        self.page.controls.append(ft.ListView([
            ft.Container(ft.Column(itens,spacing=8),padding=14)
        ],expand=True,padding=0))
        self.page.update()


def main(page: ft.Page):
    App(page)

ft.app(target=main)
